"""Bulk actions integration tests.

R2 storage operations are mocked.
"""

from unittest.mock import patch

import pytest
from httpx import ASGITransport, AsyncClient

from app.main import app


async def _register(client: AsyncClient, suffix: str = "") -> str:
    resp = await client.post(
        "/api/v1/auth/register",
        json={
            "org_name": f"Bulk Org {suffix}",
            "email": f"bulk{suffix}@test.com",
            "password": "bulkpass12345",
        },
    )
    assert resp.status_code == 201, resp.text
    token = resp.json()["access_token"]
    client.headers["Authorization"] = f"Bearer {token}"
    return token


async def _create_doc(client: AsyncClient, filename: str) -> str:
    with patch(
        "app.services.storage.presign_upload",
        return_value="https://r2.example/up",
    ):
        resp = await client.post(
            "/api/v1/documents/upload-url",
            json={
                "filename": filename,
                "content_type": "application/pdf",
                "size_bytes": 1024,
            },
        )
    assert resp.status_code == 201
    return resp.json()["document_id"]


# ---------------------------------------------------------------------------
# Bulk delete tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bulk_delete_removes_documents(client: AsyncClient) -> None:
    await _register(client, "del1")
    doc1 = await _create_doc(client, "file1.pdf")
    doc2 = await _create_doc(client, "file2.pdf")

    with patch("app.services.storage.delete_objects"):
        resp = await client.post(
            "/api/v1/documents/bulk/delete",
            json={"document_ids": [doc1, doc2]},
        )
    assert resp.status_code == 200
    assert resp.json()["deleted"] == 2

    list_resp = await client.get("/api/v1/documents")
    assert len(list_resp.json()) == 0


@pytest.mark.asyncio
async def test_bulk_delete_empty_list_returns_422(client: AsyncClient) -> None:
    await _register(client, "del2")
    resp = await client.post(
        "/api/v1/documents/bulk/delete",
        json={"document_ids": []},
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_bulk_delete_cross_org_returns_404(
    client: AsyncClient,
) -> None:
    await _register(client, "del3a")
    doc_a = await _create_doc(client, "secret.pdf")

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client_b:
        await _register(client_b, "del3b")
        with patch("app.services.storage.delete_objects"):
            resp = await client_b.post(
                "/api/v1/documents/bulk/delete",
                json={"document_ids": [doc_a]},
            )
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Bulk download tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bulk_download_returns_zip(client: AsyncClient) -> None:
    await _register(client, "dl1")
    doc1 = await _create_doc(client, "report.pdf")

    with patch(
        "app.services.storage.get_object_bytes",
        return_value=b"%PDF-1.4 fake content",
    ):
        resp = await client.post(
            "/api/v1/documents/bulk/download",
            json={"document_ids": [doc1]},
        )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"
    assert "documents-" in resp.headers["content-disposition"]
    assert len(resp.content) > 0


@pytest.mark.asyncio
async def test_bulk_download_exceeds_limit_returns_400(
    client: AsyncClient,
) -> None:
    await _register(client, "dl2")
    ids = []
    for i in range(21):
        ids.append(await _create_doc(client, f"file{i}.pdf"))

    resp = await client.post(
        "/api/v1/documents/bulk/download",
        json={"document_ids": ids},
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Bulk export tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_bulk_export_returns_xlsx(client: AsyncClient) -> None:
    await _register(client, "exp1")
    doc1 = await _create_doc(client, "invoice.pdf")

    resp = await client.post(
        "/api/v1/documents/bulk/export",
        json={"document_ids": [doc1]},
    )
    assert resp.status_code == 200
    assert (
        "spreadsheetml" in resp.headers["content-type"]
        or "application/vnd" in resp.headers["content-type"]
    )
    assert "documents-export-" in resp.headers["content-disposition"]

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert "Filename" in headers
    assert "Type" in headers
    assert "Status" in headers
    assert ws.cell(row=2, column=1).value == "invoice.pdf"


@pytest.mark.asyncio
async def test_bulk_export_cross_org_returns_404(client: AsyncClient) -> None:
    await _register(client, "exp2a")
    doc_a = await _create_doc(client, "secret.pdf")

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client_b:
        await _register(client_b, "exp2b")
        resp = await client_b.post(
            "/api/v1/documents/bulk/export",
            json={"document_ids": [doc_a]},
        )
        assert resp.status_code == 404
